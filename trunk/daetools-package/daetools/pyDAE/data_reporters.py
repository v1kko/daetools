"""********************************************************************************
                            data_reporters.py
                 DAE Tools: pyDAE module, www.daetools.com
                 Copyright (C) Dragan Nikolic
***********************************************************************************
DAE Tools is free software; you can redistribute it and/or modify it under the
terms of the GNU General Public License version 3 as published by the Free Software
Foundation. DAE Tools is distributed in the hope that it will be useful, but WITHOUT
ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS FOR A
PARTICULAR PURPOSE. See the GNU General Public License for more details.
You should have received a copy of the GNU General Public License along with the
DAE Tools software; if not, see <http://www.gnu.org/licenses/>.
********************************************************************************"""
import os, sys, numpy, json, itertools
from pyCore import *
from pyDataReporting import *

def _formatName(name):
    # Strip non alpha-numeric characters and remove the root model name.
    # I.e. _formatName('model1.model2.var') returns 'model2.var'.
    lnames = daeGetStrippedName(name).split('.')
    return '.'.join(lnames[1:])

class daeVTKDataReporter(daeDataReporterLocal):
    """
    Saves data in the binary VTK format (.vtr) using the pyEVTK module.
    pyEVTK is included in the daetools installation (daetools.ext_libs.pyevtk).
    A separate file is written into the specified directory for every time point.
    In addition, the 'variableName.visit' files are written for use with the VisIt software.
    Notate bene:
      - The original is available at https://pypi.python.org/pypi/PyEVTK. 
        Install using: "pip install pyevtk".
      - It is not an original module available at https://bitbucket.org/pauloh/pyevtk.
    Does not require VTK installed.
    """
    def __init__(self):
        daeDataReporterLocal.__init__(self)

    def Connect(self, ConnectString, ProcessName):
        """
        ConnectString is a directory where the .vtr files will be written.
        """
        self.ConnectString = ConnectString
        self.ProcessName   = ProcessName
        self.base_folder   = ConnectString
        if not os.path.isdir(self.base_folder):
            try:
                os.makedirs(self.base_folder)
            except Exception as e:
                print('Cannot create directory %s for VTK files:\n%s' % (self.base_folder, str(e)))
                return False
        
        return True

    def IsConnected(self):
        return True

    def Disconnect(self):
        self.WriteFiles()
        return True

    def WriteFiles(self):
        try:
            from daetools.ext_libs.pyevtk.hl import gridToVTK

            vtk_visit = {}
            for variable_name, (ndarr_values, ndarr_times, l_domains, s_units) in self.Process.dictVariableValues.items():
                varName = _formatName(variable_name)
                
                x = numpy.array([0.0])
                y = numpy.array([0.0])
                z = numpy.array([0.0])
                nd = len(l_domains)
                if nd > 3:
                    break

                if nd == 1:
                    x = numpy.array(l_domains[0])
                elif nd == 2:
                    x = numpy.array(l_domains[0])
                    y = numpy.array(l_domains[1])
                elif nd == 3:
                    x = numpy.array(l_domains[0])
                    y = numpy.array(l_domains[1])
                    z = numpy.array(l_domains[2])

                for (t,), time in numpy.ndenumerate(ndarr_times):
                    filename = '%s-%.5f' % (varName, time)
                    filepath = os.path.join(self.base_folder, filename)
                    if nd == 0:
                        values = numpy.array([ndarr_values[t]])
                    else:
                        values = numpy.array(ndarr_values[t])

                    values = values.reshape((len(x), len(y), len(z)))

                    gridToVTK(filepath, x, y, z, pointData = {varName : values})
                    if not varName in vtk_visit:
                        vtk_visit[varName] = []
                    vtk_visit[varName].append(filename + '.vtr')

            for var, files in vtk_visit.items():
                filename = '%s.visit' % var
                filepath = os.path.join(self.base_folder, filename)
                f = open(filepath, 'w')
                f.write('\n'.join(files))
                f.close()

        except Exception as e:
            print(('Cannot write results in .vtk format:\n' + str(e)))

class daeMatlabMATFileDataReporter(daeDataReporterFile):
    """
    Saves data in Matlab MAT format format (.mat) using scipy.io.savemat function.
    Every variable is saved as numpy array (variable names are stripped from illegal characters).
    In addition, time and domain points for every variable are saved as 'varName.Times' and 'varName.Domains'.
    Does not require Matlab installed.
    """
    def __init__(self):
        daeDataReporterFile.__init__(self)

    def WriteDataToFile(self):
        try:
            import scipy.io

            mdict = {}
            variables = self.Process.dictVariables
            for variable_name, (ndarr_values, ndarr_times, l_domains, s_units) in self.Process.dictVariableValues.items():
                name        = _formatName(variable_name)
                domainNames = [_formatName(d.Name) for d in variables[variable_name].Domains]
                
                mdict[name]                  = ndarr_values
                mdict[name + '.Times']       = ndarr_times
                mdict[name + '.DomainNames'] = numpy.array(domainNames)
                mdict[name + '.Domains']     = numpy.array(l_domains)
                mdict[name + '.Units']       = s_units

            scipy.io.savemat(self.ConnectString,
                             mdict,
                             appendmat        = True,
                             format           = '5',
                             long_field_names = True,
                             do_compression   = False,
                             oned_as          = 'row')
        except Exception as e:
            print(('Cannot write results in .mat format:\n' + str(e)))

class daeJSONFileDataReporter(daeDataReporterFile):
    """
    Saves data in JSON text format using the Python json library.
    """
    def __init__(self):
        daeDataReporterFile.__init__(self)

    def WriteDataToFile(self):
        try:
            mdict = {}
            variables = self.Process.dictVariables
            for variable_name, (ndarr_values, ndarr_times, l_domains, s_units) in self.Process.dictVariableValues.items():
                name        = _formatName(variable_name)
                domainNames = [_formatName(d.Name) for d in variables[variable_name].Domains]
                
                mdict[name] = {'Values'      : ndarr_values.tolist(),
                               'Times'       : ndarr_times.tolist(),
                               'DomainNames' : domainNames,
                               'Domains'     : [d.tolist() for d in l_domains],
                               'Units'       : s_units
                              }

            f = open(self.ConnectString, 'w')
            f.write(json.dumps(mdict, sort_keys=True, indent=4))
            f.close()
            
        except Exception as e:
            print(('Cannot write data in JSON format:\n' + str(e)))
   
class daeHDF5FileDataReporter(daeDataReporterFile):
    """
    Saves data in HDF5 format using the Python h5py library.
    A separate group is created for every variable and contain the following data sets:
     - Values: multidimensional array with the variable values
     - Times: 1d array with the time points
     - DomainNames: names of the domains that the variable is distributed on
     - Domains: multidimensional array with the domain points
     - Units: variable units as a string
    """
    def __init__(self):
        daeDataReporterFile.__init__(self)

    def WriteDataToFile(self):
        try:
            import h5py
            
            f = h5py.File(self.ConnectString, "w")
            variables = self.Process.dictVariables
            for variable_name, (ndarr_values, ndarr_times, l_domains, s_units) in self.Process.dictVariableValues.items():
                name        = _formatName(variable_name)
                domainNames = [_formatName(d.Name).encode('utf8') for d in variables[variable_name].Domains]
                
                grp = f.create_group(name)
                dsv = grp.create_dataset("Values",      data = ndarr_values)
                dst = grp.create_dataset("Times",       data = ndarr_times)
                dsd = grp.create_dataset("DomainNames", data = domainNames)
                dsd = grp.create_dataset("Domains",     data = l_domains)
                dsu = grp.create_dataset("Units",       data = s_units)

            f.close()

        except Exception as e:
            print(('Cannot write data in HDF5 format:\n' + str(e)))

class daeXMLFileDataReporter(daeDataReporterFile):
    """
    Saves data in XML format (.xml) using the Python xml library.
    The numerical data are saved as json strings (for easier parsing). 
    """
    def __init__(self):
        daeDataReporterFile.__init__(self)

    def WriteDataToFile(self):
        try:
            from xml.etree import ElementTree

            class XmlDictObject(dict):
                """
                Adds object like functionality to the standard dictionary.
                """

                def __init__(self, initdict=None):
                    if initdict is None:
                        initdict = {}
                    dict.__init__(self, initdict)

                def __getattr__(self, item):
                    return self.__getitem__(item)

                def __setattr__(self, item, value):
                    self.__setitem__(item, value)

                def __str__(self):
                    if self.has_key('_text'):
                        return self.__getitem__('_text')
                    else:
                        return ''

                @staticmethod
                def Wrap(x):
                    """
                    Static method to wrap a dictionary recursively as an XmlDictObject
                    """

                    if isinstance(x, dict):
                        return XmlDictObject((k, XmlDictObject.Wrap(v)) for (k, v) in x.items())
                    elif isinstance(x, list):
                        return [XmlDictObject.Wrap(v) for v in x]
                    else:
                        return x

                @staticmethod
                def _UnWrap(x):
                    if isinstance(x, dict):
                        return dict((k, XmlDictObject._UnWrap(v)) for (k, v) in x.items())
                    elif isinstance(x, list):
                        return [XmlDictObject._UnWrap(v) for v in x]
                    else:
                        return x

                def UnWrap(self):
                    """
                    Recursively converts an XmlDictObject to a standard dictionary and returns the result.
                    """

                    return XmlDictObject._UnWrap(self)

            def _ConvertDictToXmlRecurse(parent, dictitem):
                assert type(dictitem) is not type([])

                if isinstance(dictitem, dict):
                    for (tag, child) in dictitem.items():
                        if str(tag) == '_text':
                            parent.text = str(child)
                        elif type(child) is type([]):
                            # iterate through the array and convert
                            for listchild in child:
                                elem = ElementTree.Element(tag)
                                parent.append(elem)
                                _ConvertDictToXmlRecurse(elem, listchild)
                        else:
                            elem = ElementTree.Element(tag)
                            parent.append(elem)
                            _ConvertDictToXmlRecurse(elem, child)
                else:
                    parent.text = str(dictitem)

            def ConvertDictToXml(xmldict):
                """
                Converts a dictionary to an XML ElementTree Element
                """

                roottag = list(xmldict.keys())[0]
                root = ElementTree.Element(roottag)
                _ConvertDictToXmlRecurse(root, xmldict[roottag])
                return root

            mdict = XmlDictObject()
            xml_variables = {}
            variables = self.Process.dictVariables
            for variable_name, (ndarr_values, ndarr_times, l_domains, s_units) in sorted(self.Process.dictVariableValues.items()):
                name        = _formatName(variable_name)
                domainNames = [_formatName(d.Name) for d in variables[variable_name].Domains]

                variable = {}
                variable['Units']       = s_units
                variable['DomainNames'] = json.dumps(domainNames)
                variable['Domains']     = json.dumps(numpy.array(l_domains).tolist())
                variable['Times']       = json.dumps(ndarr_times.tolist())
                variable['Values']      = json.dumps(ndarr_values.tolist())
                xml_variables[name] = variable
            mdict['Simulation'] = xml_variables

            root = ConvertDictToXml(mdict)
            root.set('processName', self.ProcessName)
            tree = ElementTree.ElementTree(root)
            tree.write(self.ConnectString)
            
        except Exception as e:
            print(('Cannot write data in XML format:\n' + str(e)))

class daeExcelFileDataReporter(daeDataReporterFile):
    """
    Saves data into the Microsoft Excel format (.xlsx) using the openpyxl library
    (https://openpyxl.readthedocs.io).
    Does not require Excel installed and works on all operating systems.
    """
    def __init__(self):
        daeDataReporterFile.__init__(self)

    def WriteDataToFile(self):
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            for variable_name, (ndarr_values, ndarr_times, l_domains, s_units) in sorted(self.Process.dictVariableValues.items()):
                name = _formatName(variable_name)
                
                ws = wb.create_sheet(title = name)

                ws.cell(row = 1, column = 1, value = 'Time [s] / Values [%s]' % s_units)
                variable_names = ['%s%s' % (name, (list(val_indexes) if val_indexes else '')) for val_indexes, value in numpy.ndenumerate(ndarr_values[0])]
                v = 2
                for var_name in variable_names:
                    ws.cell(row = 1, column = v, value = var_name)
                    v += 1
                for (t,), time in numpy.ndenumerate(ndarr_times):
                    ws.cell(row = t+2, column = 1, value = time)
                    v = 2
                    for val_indexes, value in numpy.ndenumerate(ndarr_values[t]):
                        ws.cell(row = t+2, column = v, value = value)
                        v += 1

            wb.save(self.ConnectString)
            
        except Exception as e:
            print(('Cannot write data to an excel file:\n' + str(e)))

class daePandasDataReporter(daeDataReporterLocal):
    """
    Creates a dataset using the Pandas library 
    (available as data_frame property - the Pandas DataFrame object).
    """
    def __init__(self):
        daeDataReporterLocal.__init__(self)

    def Connect(self, ConnectString, ProcessName):
        return True

    def IsConnected(self):
        return True

    def Disconnect(self):
        return True

    @property
    def data_frame(self):
        data_frame = None
        try:
            from pandas import DataFrame
            
            names        = []
            data         = []
            domains      = []
            domain_names = []
            times        = []
            units        = []
            
            variables = self.Process.dictVariables
            for variable_name, (ndarr_values, ndarr_times, l_domains, s_units) in self.Process.dictVariableValues.items():
                name        = _formatName(variable_name)
                domainNames = [_formatName(d.Name) for d in variables[variable_name].Domains]
                
                names.append(name)
                units.append(s_units)
                times.append(ndarr_times)
                domains.append(l_domains)
                domain_names.append(domainNames)
                data.append(ndarr_values)

            _data = list(zip(units, domain_names, domains, times, data))
            data_frame = DataFrame(data = _data, columns = ['Units', 'DomainNames', 'Domains', 'Times', 'Values'], index = names)
            
        except Exception as e:
            print(('Cannot generate Pandas DataFrame:\n' + str(e)))
    
        return data_frame
    
class daePlotDataReporter(daeDataReporterLocal):
    """
    Plots the specified variables using the Matplotlib library (by Caleb Hattingh).
    """
    def __init__(self):
        daeDataReporterLocal.__init__(self)
        self.ProcessName = ""

    def Connect(self, ConnectString, ProcessName):
        self.ProcessName   = ProcessName
        self.ConnectString = ConnectString
        return True

    def IsConnected(self):
        return True

    def Disconnect(self):
        return True

    def Plot(self, *args, **kwargs):
        '''
        ``args`` can be either:

        a) Instances of daeVariable, or
        b) Lists of daeVariable instances, or
        c) A mixture of both.

        Each ``arg`` will get its own subplot. The subplots are all automatically
        arranged such that the resulting figure is as square-like as
        possible. You can however override the shape by supplying ``figRows``
        and ``figCols`` as keyword args.

        Basic Example::

            # Create Log, Solver, DataReporter and Simulation object
            log = daePythonStdOutLog()
            daesolve = daeIDAS()
            from daetools.pyDAE.data_reporters import daePlotDataReporter
            datareporter = daePlotDataReporter()
            simulation = simTutorial()

            simulation.m.SetReportingOn(True)
            simulation.ReportingInterval = 20
            simulation.TimeHorizon = 500

            simName = simulation.m.Name + strftime(" [%d.%m.%Y %H:%M:%S]", localtime())
            if(datareporter.Connect("", simName) == False):
                sys.exit()

            simulation.Initialize(daesolver, datareporter, log)

            simulation.m.SaveModelReport(simulation.m.Name + ".xml")
            simulation.m.SaveRuntimeModelReport(simulation.m.Name + "-rt.xml")

            simulation.SolveInitial()
            simulation.Run()

            simulation.Finalize()
            datareporter.Plot(
                simulation.m.Ci,                       # Subplot 1
                [simulation.m.L, simulation.m.event],  # Subplot 2 (2 sets)
                simulation.m.Vp,                       # Subplot 3
                [simulation.m.L, simulation.m.Vp]      # Subplot 4 (2 sets)
                )
        '''
        try:
            import matplotlib.pyplot as plt
            from math import sqrt, ceil, floor

            ''' Math to make sure we have enough subplots for any number
            of given variables.'''
            n = len(args)

            def kwargCheck(kw='', valueIfNone=0):
                if kw in kwargs and kwargs[kw]:
                    return kwargs[kw]
                else:
                    return valueIfNone

            rows = kwargCheck('figRows', valueIfNone=int(round(sqrt(n))))
            cols = kwargCheck('figCols', valueIfNone=int(ceil(sqrt(n))))

            ''' Create lookup for convenience.  There are two kinds of
            "variable" identifiers:
               a) instances of daeDataReceiverVariable (I call processVar)
               b) instances of daeVariable (I call daeVar)
            This lookup will create a name-value dict for later matching.'''
            lookup = {}
            for processVar in self.Process.Variables:
                lookup[processVar.Name] = processVar

            def varLabel(daeVar):
                return '{0} ({1})'.format(daeVar.Name, str(daeVar.VariableType.Units))

            def larger_axlim(axlim):
                """ This function enlarges the y-axis range so that data
                will not lie underneath an axis line.

                    axlim: 2-tuple
                    result: 2-tuple (enlarged range)

                This was taken from (thanks bernie!):
                http://stackoverflow.com/a/6230993/170656
                """
                axmin,axmax = axlim
                axrng = axmax - axmin
                new_min = axmin - 0.1 * axrng
                new_max = axmax + 0.1 * axrng
                return new_min, new_max

            def plotVariable(daeVar):
                ''' Create easy plotter that requires only a daeVariable.
                This function will plot the variable's data onto whatever
                is the current set of axes.  Note that the matchups between
                the processVar and daeVar require daeVar.CanonicalName, not
                merely Name.'''
                if daeVar.CanonicalName in lookup:
                    processVar = lookup[daeVar.CanonicalName]
                    #import pdb; pdb.set_trace()
                    plt.plot(
                        processVar.TimeValues,
                        processVar.Values,
                        label=varLabel(daeVar))

            for i, arg in enumerate(args):
                '''Loop over arguments: some are variables (each one gets its
                own subplot) and some are lists of variables, these vars
                will share a plot.'''
                axes = plt.subplot(rows, cols, i+1)
                plt.xlabel('Time (s)')
                #plt.title('Title')
                plt.grid(True)
                try:
                    plt.tight_layout(1.2)
                except Exception as e:
                    pass
                if isinstance(arg, list):
                    ''' This arg is a list of variables.  These must all be
                    plotted on the same plot.'''
                    variables = arg
                    for v in variables:
                        plotVariable(v)
                    plt.legend()
                else:
                    ''' This arg is a single variable.'''
                    variable = arg
                    plt.ylabel(varLabel(variable))
                    plotVariable(variable)
                axes.set_ylim( larger_axlim( axes.get_ylim() ) )
            plt.show()

        except Exception as e:
            import traceback
            print(('Cannot generate matplotlib plots:\n' + traceback.format_exc()))

class daeCSVFileDataReporter(daeDataReporterFile):
    """
    Saves the results in the comma-separated values (CSV) format.
    The separator is ',' and the variable names are double quoted '"'.
    """
    def __init__(self, uniqueTimeValues = False, nameBrackets = '[]'):
        daeDataReporterFile.__init__(self)

        self.uniqueTimeValues = uniqueTimeValues
        self.lBracket = nameBrackets[0]
        self.rBracket = nameBrackets[1]
        
    def WriteDataToFile(self):
        variable_names  = []
        variable_values = []
        
        max_n_times = 0
        variable_names.append('time')
        for variable_name, (ndarr_values, ndarr_times, l_domains, s_units) in self.Process.dictVariableValues.items():
            varName = _formatName(variable_name)
            domain_sizes = [range(len(d)) for d in l_domains]
            indexes = itertools.product(*domain_sizes)
            n_times = ndarr_values.shape[0]
            
            if n_times > max_n_times:
                max_n_times = n_times
                times       = ndarr_times.tolist()
                
            # Generate variable names for distributed variables
            if indexes:
                for index_list in indexes:
                    str_index_list = [str(item) for item in index_list]
                    if str_index_list:
                        inds = '%s%s%s' % (self.lBracket, ','.join(str_index_list), self.rBracket) 
                    else:
                        inds = ''
                    variable_names.append('%s%s' % (varName, inds))
                    var_indexes = [slice(0,n_times)] + list(index_list)
                    values = ndarr_values[var_indexes].tolist()
                    variable_values.append(values)
        
        # Add time values
        variable_values.insert(0, times)
        
        try:
            f = open(self.ConnectString, 'w')
            
            # Some variables (i.e. parameters) have only one value so generate values for all time points
            for i in range(len(variable_values)):
                if len(variable_values[i]) == 1:
                    variable_values[i] = variable_values[i] * max_n_times 

            # Create ndarray for easier slicing
            vals = numpy.array(variable_values, copy = False)
            
            # Write variable names in the header
            row = ','.join(['\"%s\"' % v for v in variable_names])
            f.write(row + '\n')
            
            # Write rows
            if self.uniqueTimeValues:
                unique_vals = {}
                for i in range(max_n_times):
                    row_vals = vals[..., i]
                    t = row_vals[0]
                    unique_vals[t] = row_vals
                    
                for t,row_vals in sorted(unique_vals.items()):
                    row = ','.join(['%.16e' % v for v in row_vals])
                    f.write(row + '\n')                
            else:
                for i in range(max_n_times):
                    row = ','.join(['%.16e' % v for v in vals[..., i]])
                    f.write(row + '\n')
            
            f.close()
            
        except Exception as e:
            print(('Cannot write data in CSV format:\n' + str(e)))
   
